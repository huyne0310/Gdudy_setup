﻿from tkinter import ttk
from tkinter import *
import os
import tkinter as tk
from PIL import ImageTk , Image
import time
from tkinter import messagebox
from openpyxl import load_workbook
import openpyxl
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import matplotlib.pyplot as plt

WINDOW_HEIGHT = 720
WINDOW_WIDTH = 1080
WINDOW_POSITION_LEFT = 20
WINDOW_POSITION_DOWN= 150

DIRECTORY_PATH = os.path.dirname(__file__)
PATH_IMAGES = os.path.join(DIRECTORY_PATH, "images")
PATH_DATA = os.path.join(DIRECTORY_PATH, "data")

COLOR_NUT1= '#03a1fc'

class manhinh( ):
    def __init__ (self,scr) -> None:
        scr.geometry("{}x{}+{}+{}". format(WINDOW_WIDTH,WINDOW_HEIGHT,WINDOW_POSITION_DOWN,WINDOW_POSITION_LEFT))
        scr.iconbitmap(os.path.join(PATH_IMAGES,"iconapp.ico" ))
        scr.title("Gdudy - a simple way to self-study")

scrn = tk.Tk( )
app1=manhinh(scrn)
scrn.withdraw( )
scr = Toplevel(scrn)
app2= manhinh(scr)
scr.tk_setPalette('white')
bg = PhotoImage(file = os.path.join(PATH_IMAGES, "bgi.gif"))
#grid option
scr.grid_rowconfigure(0,weight=1)
scr.grid_rowconfigure(1,weight=1)
scr.grid_rowconfigure(2,weight=1)
scr.grid_columnconfigure(0,weight=1)
scr.grid_columnconfigure(1,weight=1)
scr.grid_columnconfigure(2,weight=1)

scrn.grid_rowconfigure(0,weight=1)
scrn.grid_rowconfigure(1,weight=1)
scrn.grid_rowconfigure(2,weight=1)
scrn.grid_rowconfigure(3,weight=1)
scrn.grid_columnconfigure(0,weight=1)
scrn.grid_columnconfigure(1,weight=1)
scrn.grid_columnconfigure(2,weight=1)
scrn.grid_columnconfigure(3,weight=1)

def save_clicked_flag(value):
    with open(os.path.join(PATH_DATA,"clicked_flag.txt"), 'w',encoding='utf-8') as file:
        file.write(str(value))

def load_clicked_flag():
    try:
        with open(os.path.join(PATH_DATA,"clicked_flag.txt"), 'r') as file:
            value = file.read().strip()
            return value.lower() == 'true'
    except FileNotFoundError:
        return False
clicked_flag = load_clicked_flag()

def read_username():
    try:
        with open(os.path.join(PATH_DATA, "user_name.txt"), 'r') as file:
            NAME_USER = file.read().strip()
    except:
        pass

def childwindow():
    global clicked_flag
    if not clicked_flag:
        def clicked( ):
            NAME_USER = txt.get("1.0", "end-1c")
            with open(os.path.join(PATH_DATA, "user_name.txt"), 'w', encoding='utf-8') as file:
                file.write(NAME_USER)
            limg= Label(scr, image= bg)
            limg.place(x=0,y=0)
            #label
            lbl= Label(scr,text="Welcome to Gdudy, "+NAME_USER+" !!!", font=("Comic Sans MS",30))
            lbl.grid(row=1, column=1)
            def nextscr( ):
                limg= Label(scr, image= bg)
                limg.place(x=0,y=0)
                scr.destroy( )
                scrn.deiconify( )

            btn_frame1= Frame(scr)
            canvas = Canvas(btn_frame1, height=125, width=125)
            canvas.grid(row=1, column=1)
            btn_frame1.grid(row=2, column=1,sticky=N)
            btn1= Button(scr, text="Next", font=("Comic Sans MS",14),bg=COLOR_NUT1  ,fg= 'white', bd=0,width=8, height=1, activebackground=COLOR_NUT1, command=nextscr )
            btn1.place(x=490,y=500)
            round_rectangle(canvas, 1, 50, 125, 100, r=50, fill = COLOR_NUT1)
            
        #background image
        limg= Label(scr, image= bg)
        limg.place(x=0,y=0)
        #label
        lbl= Label(scr,text="Please enter your name:", font=("Comic Sans MS",15))
        lbl.grid(row=0, column=1,sticky=S)
        def round_rectangle(obj, x1, y1, x2, y2, r=25, **kwargs):
            points = (x1+r, y1, x1+r, y1, x2-r, y1, x2-r, y1, x2, y1, x2, y1+r, x2, y1+r, x2, y2-r, x2, y2-r, x2, y2,
                x2-r, y2, x2-r, y2, x1+r, y2, x1+r, y2, x1, y2, x1, y2-r, x1, y2-r,x1, y1+r, x1, y1+r, x1, y1)
            return obj.create_polygon(points, **kwargs, smooth = True)
        #button
        btn_frame= Frame(scr)
        canvas = Canvas(btn_frame, height=125, width=125)
        canvas.grid(row=1, column=1)
        btn_frame.grid(row=2, column=1,sticky=N)
        btn= Button(scr, text="Ok", font=("Comic Sans MS",14),bg=COLOR_NUT1  ,fg= 'white', bd=0,width=8, height=1, activebackground=COLOR_NUT1, command=clicked)
        btn.place(x=490,y=500)
        round_rectangle(canvas, 1, 50, 125, 100, r=50, fill = COLOR_NUT1)
        #textbox
        txt = Text(scr, width= 10, font=('Comic Sans MS',20), borderwidth=0, height= 3)
        txt.tag_configure("center", justify=CENTER)
        txt.tag_add("center", "1.0", "end")
        txt.insert(END, "")
        txt.grid(row=1, column=1)
        txt.focus( )
        clicked_flag = True
        save_clicked_flag(clicked_flag)
    else:
        scr.withdraw( )
        scrn.deiconify( )
        read_username( )
#close child_windows
def on_closing():
    scrn.destroy( )
scr.protocol("WM_DELETE_WINDOW", on_closing)
childwindow()

bg2 = PhotoImage(file= os.path.join(PATH_IMAGES,"scr.png"))
background2= Label(scrn,i=bg2)
background2.place(x=0,y=0)

frame2 = tk.Frame(scrn)
frame2.place(x=0,y=0)
frame2.pack_propagate(False)
frame2.configure(width= 1080, height=652)
#tabs option
tabcontrol = ttk.Notebook(frame2)
tabcontrol.pack(fill ='both',expand=True)
tab_clock = ttk.Frame(tabcontrol)
tabcontrol.add(tab_clock, text = "clock")
tab_chbx = ttk.Frame(frame2)
tabcontrol.add(tab_chbx, text = "checkbox")
tab_sch = ttk.Frame(frame2)
tabcontrol.add(tab_sch, text = "schedule")
tab_cht = ttk.Frame(frame2)
tabcontrol.add(tab_cht, text = "chart")
style = ttk.Style()
style.layout('TNotebook.Tab', [])

#change tabs
def ch2clk( ):
    tabcontrol.select(tab_clock)
def ch2sch( ):
    tabcontrol.select(tab_sch)
def ch2cht( ):
    tabcontrol.select(tab_cht)
def ch2chbx( ):
    tabcontrol.select(tab_chbx)

fclck_image=tk.PhotoImage(file= os.path.join(PATH_IMAGES, "bgclock.png"))
fclck_label = tk.Label(tab_clock, image=fclck_image)
fclck_label.place(x=0,y=0)
tab_clock.tk_setPalette('white')

#countdown clock
pomodoro = 0
h2=0  
m2=0  
s2=0
f = ("Comic Sans MS",100)
hour = StringVar()
minute = StringVar()
second = StringVar()
paused = False
if h2!=0 or m2!=0 or s2!=0:
    hour.set(h2)
    minute.set(m2)
    second.set(s2)
if h2==0 and m2==0 and s2==0:
    hour.set("00")
    minute.set("00")
    second.set("00")

def tanggio():
    global h2
    h2 = h2 + 1
    hour.set(h2)
def giamgio():
    global h2
    if h2 >0:
        h2 = h2 - 1
        hour.set(h2)
def tangphut():
    global m2
    m2=m2+1
    minute.set(m2)
def giamphut():
    global m2
    if m2>0:
        m2=m2-1
        minute.set(m2)
def tanggiay():
    global s2
    s2=s2+10
    second.set(s2)
def giamgiay():
    global s2
    if s2>0:
        s2=s2-10
        second.set(s2)
#dấu ":"
hour_tf = Entry(tab_clock, width=3, font=f, textvariable=hour,borderwidth=0,bg= 'white', disabledbackground='white')
hour_tf.place(x=168,y=180)
mins_tf = Entry(tab_clock, width=3, font=f, textvariable=minute,borderwidth=0, bg= 'white', disabledbackground='white')
mins_tf.place(x=451,y=180)
sec_tf = Entry(tab_clock, width=3, font=f, textvariable=second,borderwidth=0,bg='white', disabledbackground='white')
sec_tf.place(x=742,y=180)

def save_totaltime(f_total_time):
    with open(os.path.join(PATH_DATA, 'total_time.txt'), 'w') as f:
        f.write(str(f_total_time))

def read_totaltime():
    try:
        with open(os.path.join(PATH_DATA, 'total_time.txt'), 'r') as f:
            f_total_time = float(f.read())
        return f_total_time
    except FileNotFoundError:
        f_total_time = 0.00
        return f_total_time

def custom_messagebox( message, font):
    msgb = Toplevel(scrn)
    msgb.title("Time's up")

    msgb.geometry("500x200+300+300")
    msgb.iconbitmap(os.path.join(PATH_IMAGES,"iconapp.ico" ))
    imgtu = ImageTk.PhotoImage(Image.open(os.path.join(PATH_IMAGES, "timeup.png")))
    img_labeltu = Label(msgb, image=imgtu)
    img_labeltu.pack(side=LEFT)
    
    message_label = Label(msgb, text=message, font=font, wraplength= 300)
    message_label.pack(pady= 30 )
    
    okmsg = Image.open(os.path.join(PATH_IMAGES, "okmsg.png"))
    renderokmsg = ImageTk.PhotoImage(okmsg)
    imgokmsg = tk.Button(msgb, image=renderokmsg, borderwidth=0,highlightthickness=0,
                   highlightbackground='white',activebackground='white',command= msgb.destroy)
    imgokmsg.place(x=250, y= 130)
    msgb.resizable(False,False)
    msgb.mainloop()

clck_thuong = Label(tab_clock,text = " Clock: ", font=("Comic Sans MS", 20))
clck_thuong.place_forget( )

clck_pmdr = Label(tab_clock,text = " Pomodoro clock: ", font=("Comic Sans MS", 15))
clck_pmdr.place_forget( )

def update():
    global pomodoro
    global userinput
    global paused
    global total_time
    global h2
    global m2
    global s2
    if not paused:
        if userinput > -1:
            mins, secs = divmod(userinput, 60)
            hours = 0
        if mins > 59:
            hours, mins = divmod(mins, 60)
        hour.set("{0:2d}".format(hours))
        minute.set("{0:2d}".format(mins))
        second.set("{0:2d}".format(secs))
        clck_thuong.place(x=80, y=130)
        tab_clock.update()
        #imgpl.config(state='disabled')
        imgps.place(x=360,y=500)
        imgpl.place_forget( )
        hour_tf.bind("<Key>", lambda a: "break")
        mins_tf.bind("<Key>", lambda a: "break")
        sec_tf.bind("<Key>", lambda a: "break")
        hour_tf.config(state="readonly")
        mins_tf.config(state="readonly")
        sec_tf.config(state="readonly")
        imgmt1.config(state='disabled')
        imgmtngc2.config(state='disabled')
        imgmt3.config(state='disabled')
        imgmtngc4.config(state='disabled')
        imgmt5.config(state='disabled')
        imgmtngc6.config(state='disabled')
        imgpmdr.config(state='disabled')
        if userinput == 0:
            h2 =0
            m2=0
            s2=0
            hour_tf.unbind("<Key>")
            mins_tf.unbind("<Key>")
            sec_tf.unbind("<Key>")
            clck_thuong.place_forget( )
            custom_messagebox("Time's Up", ("Comic Sans MS",20))
            hour_tf.config(state="normal")
            mins_tf.config(state="normal")
            sec_tf.config(state="normal")
            imgmt1.config(state='normal')
            imgmtngc2.config(state='normal')
            imgmt3.config(state='normal')
            imgmtngc4.config(state='normal')
            imgmt5.config(state='normal')
            imgmtngc6.config(state='normal')
            imgpmdr.config(state='normal')
            hour.set("00")
            minute.set("00")
            second.set("00")
            imgpl.config(state="normal")
            imgpl.place(x=360,y=500)
            imgps.place_forget()
        f_total_time = read_totaltime()
        
        f_total_time = f_total_time + (1/3600)
        save_totaltime(f_total_time)
        tt_hours = Label(tab_clock,font=("Comic Sans MS",15))
        tt_hours.place(x=830, y=615)
        tt_hours.config(text = "Total time: "+ "{:.2f}".format(f_total_time)+" hours")   
        h2 = hours
        m2 = mins
        s2 = secs     
        userinput -= 1
        tab_clock.after(1000, update)


def updatepmdr():
    global pomodoro
    global userinput
    global paused
    global total_time
    global h2
    global m2
    global s2
    if not paused:
        if userinput > -1:
            mins, secs = divmod(userinput, 60)
            hours = 0
        if mins > 59:
            hours, mins = divmod(mins, 60)
        hour.set("{0:2d}".format(hours))
        minute.set("{0:2d}".format(mins))
        second.set("{0:2d}".format(secs))
        tab_clock.update()
        clck_pmdr.place(x=80, y=130)
        imgps.place(x=360,y=500)
        imgpl.place_forget( )
        imgps.place_forget( )
        hour_tf.bind("<Key>", lambda a: "break")
        mins_tf.bind("<Key>", lambda a: "break")
        sec_tf.bind("<Key>", lambda a: "break")
        hour_tf.config(state="readonly")
        mins_tf.config(state="readonly")
        sec_tf.config(state="readonly")
        imgmt1.config(state='disabled')
        imgmtngc2.config(state='disabled')
        imgmt3.config(state='disabled')
        imgmtngc4.config(state='disabled')
        imgmt5.config(state='disabled')
        imgmtngc6.config(state='disabled')
        imgpmdr.config(state='disabled')
        if userinput == 0:
            pomodoro =0
            h2 =0
            m2=0
            s2=0
            hour_tf.unbind("<Key>")
            mins_tf.unbind("<Key>")
            sec_tf.unbind("<Key>")
            clck_pmdr.place_forget( )
            custom_messagebox("Time's Up", ("Comic Sans MS",20))
            hour_tf.config(state="normal")
            mins_tf.config(state="normal")
            sec_tf.config(state="normal")
            imgmt1.config(state='normal')
            imgmtngc2.config(state='normal')
            imgmt3.config(state='normal')
            imgmtngc4.config(state='normal')
            imgmt5.config(state='normal')
            imgmtngc6.config(state='normal')
            imgpmdr.config(state='normal')
            hour.set("00")
            minute.set("00")
            second.set("00")
            imgpl.config(state="normal")
            imgpl.place(x=360,y=500)
            imgps.place_forget()
        f_total_time = read_totaltime()
        f_total_time = f_total_time + (1/3600)
        save_totaltime(f_total_time)
        tt_hours = Label(tab_clock,font=("Comic Sans MS",15))
        tt_hours.place(x=830, y=615)
        tt_hours.config(text = "Total time: "+ "{:.2f}".format(f_total_time)+" hours")   
        h2 = hours
        m2 = mins
        s2 = secs     
        userinput -= 1
        pomodoro += 1
        if pomodoro % 6000 == 0:
            tab_clock.after(900000, updatepmdr)
            custom_messagebox(" Hãy nghỉ ngơi 15 phút rồi tiếp tục nhé! ", ("Arial",20))
        elif pomodoro % 1500 == 0:
            tab_clock.after(300000, updatepmdr)
            custom_messagebox(" Hãy nghỉ ngơi 5 phút rồi tiếp tục nhé! ", ("Arial",20))
        else:
            tab_clock.after(1000, updatepmdr)

def startpmdr():
    global userinput   
    global paused
    paused = False
    try:
        userinput = int(hour.get()) * 3600 + int(minute.get()) * 60 + int(second.get())
        if userinput == 0:
            userinput = h2*3600 + m2*60+ s2
    except:
        messagebox.showwarning('Invalid Input!')
    tab_clock.update()
    updatepmdr( )

def startcd():
    global userinput   
    global paused
    paused = False
    try:
        userinput = int(hour.get()) * 3600 + int(minute.get()) * 60 + int(second.get())
        if userinput == 0:
            userinput = h2*3600 + m2*60+ s2
    except:
        messagebox.showwarning('Invalid Input!')
    tab_clock.update()
    update( )

def muiten_time():
    global h2
    global m2
    global s2
    startcd( )
    
def enable_btnrs():
    imgrs.config(state="normal")
def enable_btnps():
    imgps.config(state="normal")
def pause_resume( ):
        global userinput
        global paused
        paused = not paused
        imgmt1.config(state='normal')
        imgmtngc2.config(state='normal')
        imgmt3.config(state='normal')
        imgmtngc4.config(state='normal')
        imgmt5.config(state='normal')
        imgmtngc6.config(state='normal')        
        def enable_btnrs():
            imgrs.config(state="normal")
        def enable_btnps():
            imgps.config(state="normal")
        def dis_btnrs():
            imgrs.config(state="disabled")
            tab_clock.after(1000, enable_btnrs)
        def dis_btnps():
            imgps.config(state="disabled")
            tab_clock.after(1000, enable_btnps)
        imgps.place_forget()
        imgrs.place(x=360,y=500)
        if not paused:
            muiten_time()
        if paused == False:
            imgrs.place_forget()
            imgps.place(x=360,y=500)
            dis_btnps()
        if paused == True:
            imgps.place_forget()
            imgrs.place(x=360,y=500)
            dis_btnrs()

def resetcd():
    global userinput
    global paused
    global h2
    global s2
    global m2
    global pomodoro
    h2=0
    s2=0
    m2=0
    pomodoro = 0
    userinput = 0
    paused = True
    imgpmdr.config(state='normal')
    hour_tf.unbind("<Key>")
    mins_tf.unbind("<Key>")
    sec_tf.unbind("<Key>")
    imgmt1.config(state='normal')
    imgmtngc2.config(state='normal')
    imgmt3.config(state='normal')
    imgmtngc4.config(state='normal')
    imgmt5.config(state='normal')
    imgmtngc6.config(state='normal')    
    hour_tf.config(state="normal")
    mins_tf.config(state="normal")
    sec_tf.config(state="normal")
    clck_pmdr.place_forget( )
    clck_thuong.place_forget( )
    imgps.place_forget()
    imgrs.place_forget()
    imgpl.place(x=360,y=500)
    hour.set("00")
    minute.set("00")
    second.set("00")
    tab_clock.update()
    
clccdlbl1 = Label(tab_clock, text=":",font=f,fg='black')
clccdlbl1.place(x=404,y=167)
clccdlbl2 = Label(tab_clock, text=":",font=f,fg='black')
clccdlbl2.place(x=696,y=167)

play = Image.open(os.path.join(PATH_IMAGES, "playbtn.png"))
render01 = ImageTk.PhotoImage(play)
imgpl = tk.Button(tab_clock, image=render01, borderwidth=0,command= startcd)
imgpl.image = render01
imgpl.place(x=360,y=500)
pause = Image.open(os.path.join(PATH_IMAGES, "pausebtn.png"))
render02 = ImageTk.PhotoImage(pause)
imgps = tk.Button(tab_clock, image=render02, borderwidth=0,command= pause_resume)
imgps.image = render02
#imgps.place(x=360,y=500) #x=485
imgps.place_forget()
reset = Image.open(os.path.join(PATH_IMAGES, "resetbtn.png"))
render03 = ImageTk.PhotoImage(reset)
imgrs = tk.Button(tab_clock, image=render03, borderwidth=0,command=resetcd)
imgrs.image = render03
imgrs.place(x=610,y=500)
resume = Image.open(os.path.join(PATH_IMAGES, "playbtn.png"))
render04 = ImageTk.PhotoImage(resume)
imgrs = tk.Button(tab_clock, image=render04, borderwidth=0,command= pause_resume)
imgrs.image = render04

pmdr_btn = Image.open(os.path.join(PATH_IMAGES, "pmdr.png"))
renderpmdr = ImageTk.PhotoImage(pmdr_btn)
imgpmdr = tk.Button(tab_clock, image=renderpmdr, borderwidth=0, command= startpmdr)
imgpmdr.image = renderpmdr
imgpmdr.place(x=485, y=500)
#imgps.place(x=360,y=500) #x=485
imgrs.place_forget()

#mui ten
#1-2: giay
mt1 = Image.open(os.path.join(PATH_IMAGES, "muiten.png"))
rendermt1 = ImageTk.PhotoImage(mt1)
imgmt1 = tk.Button(tab_clock, image=rendermt1, borderwidth=0, highlightcolor='white',highlightthickness=0,
                   highlightbackground='white',activebackground='white',command= tanggiay)
imgmt1.image = rendermt1
imgmt1.place(x=778, y=165)
#3-4: phut
mt3 = Image.open(os.path.join(PATH_IMAGES, "muiten.png"))
rendermt3 = ImageTk.PhotoImage(mt3)
imgmt3 = tk.Button(tab_clock, image=rendermt3, borderwidth=0, highlightcolor='white',highlightthickness=0,
                   highlightbackground='white',activebackground='white', command=tangphut)
imgmt3.image = rendermt3
imgmt3.place(x=490, y=165)
#5-6: gio
mt5 = Image.open(os.path.join(PATH_IMAGES, "muiten.png"))
rendermt5 = ImageTk.PhotoImage(mt5)
imgmt5 = tk.Button(tab_clock, image=rendermt3, borderwidth=0, highlightcolor='white',highlightthickness=0,
                   highlightbackground='white',activebackground='white', command= tanggio)
imgmt5.image = rendermt5
imgmt5.place(x=210, y=165)

mt2 = Image.open(os.path.join(PATH_IMAGES, "muiten.png"))
mtngc2 = mt2.transpose(Image.FLIP_TOP_BOTTOM)
rendermtngc2 = ImageTk.PhotoImage(mtngc2)
imgmtngc2 = tk.Button(tab_clock, image=rendermtngc2, borderwidth=0, highlightcolor='white',highlightthickness=0,
                      highlightbackground='white',activebackground='white', command= giamgiay)
imgmtngc2.image = rendermtngc2
imgmtngc2.place(x=778, y=345)

mt4 = Image.open(os.path.join(PATH_IMAGES, "muiten.png"))
mtngc4 = mt4.transpose(Image.FLIP_TOP_BOTTOM)
rendermtngc4 = ImageTk.PhotoImage(mtngc4)
imgmtngc4 = tk.Button(tab_clock, image=rendermtngc4, borderwidth=0, highlightcolor='white',highlightthickness=0,
                      highlightbackground='white',activebackground='white',command=giamphut)
imgmtngc4.image = rendermtngc4
imgmtngc4.place(x=490, y=345)

mt6 = Image.open(os.path.join(PATH_IMAGES, "muiten.png"))
mtngc6 = mt6.transpose(Image.FLIP_TOP_BOTTOM)
rendermtngc6 = ImageTk.PhotoImage(mtngc6)
imgmtngc6 = tk.Button(tab_clock, image=rendermtngc6, borderwidth=0, highlightcolor='white',highlightthickness=0,
                      highlightbackground='white',activebackground='white', command= giamgio)
imgmtngc6.image = rendermtngc6
imgmtngc6.place(x=210, y=345)

#button app
clock = Image.open(os.path.join(PATH_IMAGES, "clock-remaining.png"))
render1 = ImageTk.PhotoImage(clock)
imgclk = tk.Button(scrn, image=render1, borderwidth=0,command=ch2clk, highlightcolor='#E8DBD0',highlightthickness=0,highlightbackground='#E8DBD0',activebackground='#DFD0C2')
imgclk.image = render1
imgclk.grid(row = 3, column=2, sticky= S, pady=5)

chbx = Image.open(os.path.join(PATH_IMAGES, "chbx.png"))
render2 = ImageTk.PhotoImage(chbx)
imgchb = tk.Button(scrn, image=render2, borderwidth=0,command=ch2chbx, highlightcolor='#E8DBD0',highlightthickness=0,highlightbackground='#E8DBD0',activebackground='#DFD0C2')
imgchb.image = render2
imgchb.grid(row = 3, column=1, sticky= S,pady=5)

sch = Image.open(os.path.join(PATH_IMAGES, "schedule.png"))
render3 = ImageTk.PhotoImage(sch)
imgsch = tk.Button(scrn, image=render3, borderwidth=0,command=ch2sch, highlightcolor='#E8DBD0',highlightthickness=0,highlightbackground='#E8DBD0',activebackground='#DFD0C2')
imgsch.image = render3
imgsch.grid(row = 3, column=0, sticky= S,pady=5)

cht = Image.open(os.path.join(PATH_IMAGES, "fan-chart.png"))
render4 = ImageTk.PhotoImage(cht)
imgcht = tk.Button(scrn, image=render4, borderwidth=0,command=ch2cht, highlightcolor='#E8DBD0',highlightthickness=0,highlightbackground='#E8DBD0',activebackground='#DFD0C2')
imgcht.image = render4
imgcht.grid(row = 3, column=3, sticky= S,pady=5)

#tab-chbx
frame3 = tk.Frame(tab_chbx, bg = 'white', width=1080, height= 653)
frame3.place(x=0,y=0)
frame3.pack_propagate(False)

bgchbx = PhotoImage(file= os.path.join(PATH_IMAGES,"bgchbx.png"))
backgroundchbx= Label(tab_chbx,i=bgchbx)
backgroundchbx.propagate(False)
backgroundchbx.place(x=0,y=0)

frame4 = tk.Frame(tab_chbx, bg= 'white', height=345, width=775)
frame4.place(x=155,y=145)
frame4.pack_propagate(False)

addtkslbl = Label(tab_chbx, text= "Please enter your tasks: ", font=('Comic Sans MS',14))
addtkslbl.place(x=10,y=548)
addtasks_box = Entry(tab_chbx, font=("Arial",20),width=56, borderwidth=0,bg='white', highlightthickness=0.5,highlightcolor='white',highlightbackground='white')
addtasks_box.place(x=25,y=595)

tasks = []

#tạo theme cho checkbox hình trònnn....(bing AI)
style = ttk.Style()
image_on = Image.open( os.path.join(PATH_IMAGES,"chbxon.png"))
image_off = Image.open(os.path.join(PATH_IMAGES,"chbxoff.png"))
image_on = ImageTk.PhotoImage(image_on)
image_off = ImageTk.PhotoImage(image_off)
style.element_create("custom.Checkbutton.indicator", "image", image_off,
                     ("selected", image_on), ("active", image_on),
                     width=image_on.width(), sticky="w")
style.layout("CustomCheckbutton",
             [("custom.Checkbutton.padding",
               {"sticky": "nswe",
                "children": [("custom.Checkbutton.indicator", {"side": "left", "sticky": ""}),
                             ("custom.Checkbutton.label", {"sticky": "nswe"})]})])
style.configure("CustomCheckbutton", background="white")
style.configure("CustomCheckbutton", font=("Arial", 14))

def update_checkboxes():
    for widget in frame4.winfo_children():
        widget.destroy()
    
    canvas = tk.Canvas(frame4)
    scrollbar = tk.Scrollbar(frame4, orient="vertical", command=canvas.yview, width=7)
    if canvas.winfo_exists():
        canvas.configure(highlightthickness=0, highlightbackground="white")
    scrollable_frame = tk.Frame(canvas)
    def on_configure(event):
        canvas.configure(scrollregion=canvas.bbox("all"))
        if scrollable_frame.winfo_height() > canvas.winfo_height():
            scrollbar.pack(side="right", fill="y")
            canvas.bind("<MouseWheel>", on_mousewheel)
        else:
            scrollbar.pack_forget()
            canvas.unbind("<MouseWheel>")

    scrollable_frame.bind("<Configure>", on_configure)

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)

    def on_checkbutton_select(task):
        tasks.remove(task)
        with open(os.path.join(PATH_DATA, "tasks.txt"), "w", encoding = "utf-8") as f:
            for task in tasks:
                f.write(task + "\n")
        update_checkboxes()
        if canvas.winfo_exists():
            canvas.configure(highlightthickness=0, highlightbackground="white")

    for task in tasks:
        var = tk.IntVar()
        frame = tk.Frame(scrollable_frame)
        c = ttk.Checkbutton(frame, variable=var, style="CustomCheckbutton", command=lambda task=task: on_checkbutton_select(task))
        c.pack(side=tk.LEFT)
        l = tk.Label(frame, text=task, wraplength=725, font=('Arial',14))
        l.pack(side=tk.LEFT)
        frame.pack(side=tk.TOP, anchor=tk.W, pady=(10,0))

    def on_mousewheel(event):
        canvas.yview_scroll(int(-1*(event.delta/120)), "units")

    scrn.bind("<MouseWheel>", on_mousewheel)

    canvas.pack(side="left", fill="both", expand=True)

def save_tasks(event = None):
    task = addtasks_box.get()
    if task != "":
        tasks.append(task)
        update_checkboxes()
    else:
        messagebox.showwarning("Warning", "Please enter a task.")
    addtasks_box.delete(0, tk.END)
    with open(os.path.join(PATH_DATA, "tasks.txt"), "w", encoding = "utf-8") as f:
        for task in tasks:
            f.write(task + "\n")
def load_tasks():
    try:
        with open(os.path.join(PATH_DATA, "tasks.txt"), "r", encoding = "utf-8") as f:
            lines = f.readlines()
            for line in lines:
                tasks.append(line.strip())
    except FileNotFoundError:
        pass

addtask_btn = Image.open(os.path.join(PATH_IMAGES, "addtasks.png"))
render5 = ImageTk.PhotoImage(addtask_btn)
imgadtk = tk.Button(tab_chbx, image=render5, borderwidth=0, command= save_tasks)
imgadtk.image = render5
imgadtk.place(x=940,y=525)

addtasks_box.bind('<Return>', save_tasks)

gdutks = Label(tab_chbx, text="Your tasks", font=("Comic Sans MS", 30))
gdutks.place(x=315,y=15)

load_tasks()
update_checkboxes()

bgsch = tk.PhotoImage(file=os.path.join(PATH_IMAGES, "bgschedule.png"))
backgroundsch = tk.Label(tab_sch, image=bgsch)
backgroundsch.propagate(False)
backgroundsch.place(x=0, y=0)

#Monday
txtmonday = tk.Text(tab_sch, font=("Arial", 20), height=5, width=11, highlightcolor='white', borderwidth=0)
txtmonday.place(x=65, y=80)
def save_datat2():
    with open(os.path.join(PATH_DATA, "monday.txt"), "w", encoding='utf-8') as file:
        file.write(txtmonday.get("1.0", tk.END))
def load_monday():
    try:
        with open(os.path.join(PATH_DATA, "monday.txt"), "r", encoding="utf-8") as f:
            lines = f.readlines()
            txtmonday.delete("1.0", tk.END)
            for line in lines:
                txtmonday.insert(tk.END, line)
    except FileNotFoundError:
        pass
load_monday()

#Tuesday
txttuesday = tk.Text(tab_sch, font=("Arial", 20), height=5, width=11, highlightcolor='white', borderwidth=0)
txttuesday.place(x=320, y=148)
def save_datat3():
    with open(os.path.join(PATH_DATA, "tuesday.txt"), "w", encoding='utf-8') as file:
        file.write(txttuesday.get("1.0", tk.END))
def load_tuesday():
    try:
        with open(os.path.join(PATH_DATA, "tuesday.txt"), "r", encoding="utf-8") as f:
            lines = f.readlines()
            txttuesday.delete("1.0", tk.END)
            for line in lines:
                txttuesday.insert(tk.END, line)
    except FileNotFoundError:
        pass
load_tuesday()

#Wednesday
txtwed = tk.Text(tab_sch, font=("Arial", 20), height=5, width=11, highlightcolor='white', borderwidth=0)
txtwed.place(x=590, y=99)
def save_datat4():
    with open(os.path.join(PATH_DATA, "wednesday.txt"), "w", encoding='utf-8') as file:
        file.write(txtwed.get("1.0", tk.END))
def load_wed():
    try:
        with open(os.path.join(PATH_DATA, "wednesday.txt"), "r", encoding="utf-8") as f:
            lines = f.readlines()
            txtwed.delete("1.0", tk.END)
            for line in lines:
                txtwed.insert(tk.END, line)
    except FileNotFoundError:
        pass
load_wed()

#Thursday
txtthr = tk.Text(tab_sch, font=("Arial", 20), height=5, width=11, highlightcolor='white', borderwidth=0)
txtthr.place(x=852, y=149)
def save_datat5():
    with open(os.path.join(PATH_DATA, "thursday.txt"), "w", encoding='utf-8') as file:
        file.write(txtthr.get("1.0", tk.END))
def load_thr():
    try:
        with open(os.path.join(PATH_DATA, "thursday.txt"), "r", encoding="utf-8") as f:
            lines = f.readlines()
            txtthr.delete("1.0", tk.END)
            for line in lines:
                txtthr.insert(tk.END, line)
    except FileNotFoundError:
        pass
load_thr()

#Friday
txtfri = tk.Text(tab_sch, font=("Arial", 20), height=5, width=11, highlightcolor='white', borderwidth=0)
txtfri.place(x=123, y=409)
def save_datat6():
    with open(os.path.join(PATH_DATA, "friday.txt"), "w", encoding='utf-8') as file:
        file.write(txtfri.get("1.0", tk.END))
def load_fri():
    try:
        with open(os.path.join(PATH_DATA, "friday.txt"), "r", encoding="utf-8") as f:
            lines = f.readlines()
            txtfri.delete("1.0", tk.END)
            for line in lines:
                txtfri.insert(tk.END, line)
    except FileNotFoundError:
        pass
load_fri()

#Saturday
txtstrd = tk.Text(tab_sch, font=("Arial", 20), height=5, width=11, highlightcolor='white', borderwidth=0)
txtstrd.place(x=457, y=452)
def save_datat7():
    with open(os.path.join(PATH_DATA, "saturday.txt"), "w", encoding='utf-8') as file:
        file.write(txtstrd.get("1.0", tk.END))
def load_strd():
    try:
        with open(os.path.join(PATH_DATA, "saturday.txt"), "r", encoding="utf-8") as f:
            lines = f.readlines()
            txtstrd.delete("1.0", tk.END)
            for line in lines:
                txtstrd.insert(tk.END, line)
    except FileNotFoundError:
        pass
load_strd()

#Sunday
txtsd = tk.Text(tab_sch, font=("Arial", 20), height=5, width=11, highlightcolor='white', borderwidth=0)
txtsd.place(x=784, y=449)
def save_datacn():
    with open(os.path.join(PATH_DATA, "sunday.txt"), "w", encoding='utf-8') as file:
        file.write(txtsd.get("1.0", tk.END))
def load_sd():
    try:
        with open(os.path.join(PATH_DATA, "sunday.txt"), "r", encoding="utf-8") as f:
            lines = f.readlines()
            txtsd.delete("1.0", tk.END)
            for line in lines:
                txtsd.insert(tk.END, line)
    except FileNotFoundError:
        pass
load_sd()

def save_button_on():
    save_datat2()
    save_datat3()
    save_datat4()
    save_datat5()
    save_datat6()
    save_datat7()
    save_datacn()
savesch_btn = Image.open(os.path.join(PATH_IMAGES, "savebtn.png"))
render6 = ImageTk.PhotoImage(savesch_btn)
imgsvsch = tk.Button(tab_sch, image=render6, borderwidth=0, command= save_button_on,highlightthickness=0, highlightcolor='#EADED4',highlightbackground='#EADED4')
imgsvsch.image = render6
imgsvsch.place(x=573,y=295)

#menubox
def menu_box():
    def cl_mn():
        menu_frame.place_forget( )
        imgmn.config(image= rendermn)
        imgmn.config(command= menu_box)

    menu_frame = Frame(frame2, width= 170, height=230, bg='#EADED4')
    menu_frame.place(x=900,y=15)

    clock_btn = Button(menu_frame,text="G-clock", font=("Comic Sans MS",14), bg='#EADED4', borderwidth=0, command= ch2clk, activebackground='#EADED4')
    clock_btn.pack()
    chbx_btn = Button(menu_frame,text="G-tasks", font=("Comic Sans MS",14), bg='#EADED4', borderwidth=0, command= ch2chbx, activebackground='#EADED4')
    chbx_btn.pack()
    sch_btn = Button(menu_frame,text="G-schedule", font=("Comic Sans MS",14), bg='#EADED4', borderwidth=0, command= ch2sch, activebackground='#EADED4')
    sch_btn.pack()
    cht_btn = Button(menu_frame,text="G-charts", font=("Comic Sans MS",14), bg='#EADED4', borderwidth=0, command= ch2cht, activebackground='#EADED4')
    cht_btn.pack()
    imgmn.config(image= renderclmn)
    imgmn.config(command= cl_mn)


menu = Image.open(os.path.join(PATH_IMAGES, "menu.png"))
rendermn = ImageTk.PhotoImage(menu)
imgmn = tk.Button(frame2, image=rendermn, borderwidth=0,highlightthickness=0,
                  highlightcolor='#EADED4',highlightbackground='#EADED4', command= menu_box)
imgmn.image = rendermn
imgmn.place(x=1010,y=15)

clmenu = Image.open(os.path.join(PATH_IMAGES, "clmenu.png"))
renderclmn = ImageTk.PhotoImage(clmenu)

#AI
aifri = tk.PhotoImage(file=os.path.join(PATH_IMAGES, "AIframe.png"))
aifr1 = tk.Label(tab_clock, image=aifri, highlightthickness=0,highlightcolor='#9a9b9c',highlightbackground='#9a9b9c')
aifr1.propagate(False)
aifr1.place(x=780, y=453)

def hide_aifr1():
    aifr1.place_forget( )    
chao = Label(aifr1, text="Chào mừng bạn đã đến với Gdudy. Tôi rất mong bạn sẽ có một trải nghiệm tuyệt vời nhất!", 
            wraplength=220,font=("Arial",14))
chao.pack()
allow_btnimg = Image.open(os.path.join(PATH_IMAGES, "allowbtn.png"))
render7 = ImageTk.PhotoImage(allow_btnimg)
imgallow = tk.Button(aifr1, image=render7, borderwidth=0,highlightthickness=0,
                        highlightcolor='white',highlightbackground='white',command=hide_aifr1)
imgallow.image = render7
imgallow.place(x=120,y=115)
aifr1.after(8000,hide_aifr1)

#charts
bg_cht = PhotoImage(file = os.path.join(PATH_IMAGES, "bg_cht.png"))
bg_chtl = Label(tab_cht,image= bg_cht)
bg_chtl.place(x=0, y=0)
bg_chtl.propagate(False)

#ve bieu do

def vebd():
    work_book = openpyxl.load_workbook(os.path.join(PATH_DATA, "diem.xlsx"))
    sheet = work_book.active
    pcht=[]
    pcht_mon = []

    MAU1 = '#b3f2f2'
    MAU2 = '#bce0b6'
    MAU3 = '#afd67c'
    MAU4 = '#fce17e'
    MAU5 = '#fae0b4'
    MAU6 = '#fabdb4'
    MAU7 = '#fab4dd'
    MAU8 = '#b4b5fa'
    MAU9 = '#68beed'
    MAU10 = '#68ede0'
    MAU11 = '#6ced68'
    MAU12 = '#ed8e68'
    mau_btd = [MAU1,MAU2,MAU3,MAU4,MAU5,MAU6,MAU7,MAU8,MAU9,MAU10,MAU11,MAU12]
    mau_ve=[]
    def ip_bdtron():
        tbt = sheet.cell(row = 8, column = 1)
        if tbt.value is not None:
            pcht.append(float(tbt.value))
        else:
            pass
        tbv = sheet.cell(row = 8, column = 2)
        if tbv.value is not None:
            pcht.append(float(tbv.value))
        else:
            pass
        tba = sheet.cell(row = 8, column = 3)
        if tba.value is not None:
            pcht.append(float(tba.value))
        else:
            pass
        tbqp = sheet.cell(row = 8, column = 4)
        if tbqp.value is not None:
            pcht.append(float(tbqp.value))
        else:
            pass
        tbls = sheet.cell(row = 8, column = 5)
        if tbls.value is not None:
            pcht.append(float(tbls.value))
        else:
            pass
        tbdl = sheet.cell(row = 8, column = 6)
        if tbdl.value is not None:
            pcht.append(float(tbdl.value))
        else:
            pass
        tbkt = sheet.cell(row = 8, column = 7)
        if tbkt.value is not None:
            pcht.append(float(tbkt.value))
        else:
            pass
        tbti = sheet.cell(row = 8, column = 8)
        if tbti.value is not None:
            pcht.append(float(tbti.value))
        else:
            pass
        tbvl = sheet.cell(row = 8, column = 9)
        if tbvl.value is not None:
            pcht.append(float(tbvl.value))
        else:
            pass
        tbh = sheet.cell(row = 8, column = 10)
        if tbh.value is not None:
            pcht.append(float(tbh.value))
        else:
            pass
        tbs = sheet.cell(row = 8, column = 11)
        if tbs.value is not None:
            pcht.append(float(tbs.value))
        else:
            pass

    def mon_bdtron():
        for m in range(1,12):
            t = sheet.cell(row = 1, column = m)
            k = sheet.cell(row = 8, column = m)
            if k.value is not None:
                pcht_mon.append(str(t.value))
            else:
                pass


    mon_bdtron( )
    ip_bdtron( )
    font_bdt = {'family': 'Arial', 'size': 7}
    kc_bdtron = []
    for i in range (len(pcht_mon)):
        kc_bdtron.append(0.01)

    for l in range (len(pcht_mon)):
        mau_ve.append(mau_btd[l])

    figpcht = Figure(figsize=(3, 3))
    axpcht = figpcht.add_subplot(111)
    axpcht.pie(pcht, labels= pcht_mon, explode=kc_bdtron, textprops= font_bdt, colors = mau_ve)
    bdotron = FigureCanvasTkAgg(figpcht, master= tab_cht)
    bdotron.draw()
    bdotron.get_tk_widget().place(x=650,y=145)

    #bieu do cot
    figbdc, axbdc = plt.subplots(figsize = (5,4))
    axbdc.bar(pcht_mon, pcht, color = '#75b2eb')
    font_bdc = {'family': 'Arial', 'size': 7}
    axbdc.tick_params(axis='x', labelsize=10)
    axbdc.set_xticks(range(len(pcht_mon)))
    axbdc.set_xticklabels(pcht_mon, fontdict=font_bdc)
    bdoc = FigureCanvasTkAgg(figbdc, master= tab_cht)
    bdoc.draw()
    bdoc.get_tk_widget().place(x=50, y=150)

def cht_ip():
    cht_input = Toplevel(scrn)
    cht_input.geometry('950x650')
    cht_input.title ("Vui lòng nhập điểm của bạn:")
    cht_input.iconbitmap(os.path.join(PATH_IMAGES,"iconapp.ico" ))
    ipbg = PhotoImage(file= os.path.join(PATH_IMAGES, "bgcht_ip.png"))
    bg_ip = Label(cht_input, image= ipbg)
    bg_ip.place(x=0,y=0)

    def save_diem(diem,o ):
        wb = load_workbook(os.path.join(PATH_DATA, "diem.xlsx"))
        ws = wb['Diem']
        ws[o] = diem
        return wb.save(os.path.join(PATH_DATA, "diem.xlsx"))
    def load_diem(o, cho):
        wb = load_workbook(os.path.join(PATH_DATA, "diem.xlsx"))
        ws = wb['Diem']
        value = ws[o].value
        cho.delete(0, 'end')
        cho.insert(0, value)
        
    #toan
    tx1t = Entry(cht_input, width=3, font = ("Comic Sans MS", 13), borderwidth=0)
    tx1t.place(x=32,y=105)

    tx2t = Entry(cht_input, width=3, font = ("Comic Sans MS", 13), borderwidth=0)
    tx2t.place(x=69,y=105)
    tx3t = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx3t.place(x=109,y=105)
    tx4t = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx4t.place(x=148,y=105)
    txgkt = Entry(cht_input, width=6 , font = ("Comic Sans MS",13), borderwidth=0)
    txgkt.place(x=40,y=171)
    txckt = Entry(cht_input, width=6 , font = ("Comic Sans MS",13), borderwidth=0)
    txckt.place(x=109,y=171)

    def diem_toan():
        save_diem(tx1t.get(),'A2')
        save_diem(tx2t.get(),'A3')
        save_diem(tx3t.get(),'A4')
        save_diem(tx4t.get(),'A5')
        save_diem(txgkt.get(),'A6')
        save_diem(txckt.get(),'A7')
        try:
            diem_tbt = (float(tx1t.get()) + float(tx2t.get()) + float(tx3t.get()) + float(tx4t.get())
                        + float((txgkt.get()))*2 + float((txckt.get()))*3)/9
            save_diem(diem_tbt, 'A8')
        except:
            pass

    #van
    tx1v = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx1v.place(x=267,y=105)
    tx2v = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx2v.place(x=304,y=105)
    tx3v = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx3v.place(x=346,y=105)
    tx4v = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx4v.place(x=384,y=105)
    txgkv = Entry(cht_input, width=6 ,  font = ("Comic Sans MS",13), borderwidth=0)
    txgkv.place(x=276,y=171)
    txckv = Entry(cht_input, width=6 ,  font = ("Comic Sans MS",13), borderwidth=0)
    txckv.place(x=345,y=171)

    def diem_van():
        save_diem(tx1v.get(),'B2')
        save_diem(tx2v.get(),'B3')
        save_diem(tx3v.get(),'B4')
        save_diem(tx4v.get(),'B5')
        save_diem(txgkv.get(),'B6')
        save_diem(txckv.get(),'B7')
        try:
            diem_tbv = (float(tx1v.get()) + float(tx2v.get()) + float(tx3v.get()) + float(tx4v.get())
                        + float((txgkv.get()))*2 + float((txckv.get()))*3)/9
            save_diem(diem_tbv, 'B8')
        except:
            pass
    #anh
    tx1a = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx1a.place(x=499,y=105)    
    tx2a = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx2a.place(x=538,y=105)    
    tx3a = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx3a.place(x=577,y=105)    
    tx4a = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx4a.place(x=614,y=105)
    txgka = Entry(cht_input, width=6 ,  font = ("Comic Sans MS",13), borderwidth=0)
    txgka.place(x=510,y=170)
    txcka = Entry(cht_input, width=6 ,  font = ("Comic Sans MS",13), borderwidth=0)
    txcka.place(x=576,y=170)

    def diem_anh():
        save_diem(tx1a.get(),'C2')
        save_diem(tx2a.get(),'C3')
        save_diem(tx3a.get(),'C4')
        save_diem(tx4a.get(),'C5')
        save_diem(txgka.get(),'C6')
        save_diem(txcka.get(),'C7')
        try:
            diem_tba = (float(tx1a.get()) + float(tx2a.get()) + float(tx3a.get()) + float(tx4a.get())
                        + float((txgka.get()))*2 + float((txcka.get()))*3)/9
            save_diem(diem_tba, 'C8')
        except:
            pass
    #GDQP
    tx1qp = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx1qp.place(x=731,y=105)        
    tx2qp = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx2qp.place(x=767,y=105) 
    tx3qp = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx3qp.place(x=807,y=105) 
    tx4qp = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx4qp.place(x=845,y=105)
    txgkqp = Entry(cht_input, width=6 ,  font = ("Comic Sans MS",13), borderwidth=0)
    txgkqp.place(x=737,y=170)
    txckqp = Entry(cht_input, width=6 ,  font = ("Comic Sans MS",13), borderwidth=0)
    txckqp.place(x=807,y=170) 

    def diem_gdqp():
        save_diem(tx1qp.get(),'D2')
        save_diem(tx2qp.get(),'D3')
        save_diem(tx3qp.get(),'D4')
        save_diem(tx4qp.get(),'D5')
        save_diem(txgkqp.get(),'D6')
        save_diem(txckqp.get(),'D7')
        try:
            diem_tbqp = (float(tx1qp.get()) + float(tx2qp.get()) + float(tx3qp.get()) + float(tx4qp.get())
                        + float((txgkqp.get()))*2 + float((txckqp.get()))*3)/9
            save_diem(diem_tbqp, 'D8')
        except:
            pass

    #ls
    tx1ls = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx1ls.place(x=72,y=313)        
    tx2ls = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx2ls.place(x=109,y=313) 
    tx3ls = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx3ls.place(x=152,y=313) 
    tx4ls = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx4ls.place(x=189,y=313)
    txgkls = Entry(cht_input, width=6 ,  font = ("Comic Sans MS",13), borderwidth=0)
    txgkls.place(x=78,y=382)
    txckls = Entry(cht_input, width=6 ,  font = ("Comic Sans MS",13), borderwidth=0)
    txckls.place(x=150,y=382)     

    def diem_su():
        save_diem(tx1ls.get(),'E2')
        save_diem(tx2ls.get(),'E3')
        save_diem(tx3ls.get(),'E4')
        save_diem(tx4ls.get(),'E5')
        save_diem(txgkls.get(),'E6')
        save_diem(txckls.get(),'E7')
        try:
            diem_tbls = (float(tx1ls.get()) + float(tx2ls.get()) + float(tx3ls.get()) + float(tx4ls.get())
                        + float((txgkls.get()))*2 + float((txckls.get()))*3)/9
            save_diem(diem_tbls, 'E8')
        except:
            pass

    #Địa
    tx1dl = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx1dl.place(x=314,y=313)        
    tx2dl = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx2dl.place(x=350,y=313) 
    tx3dl = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx3dl.place(x=391,y=313) 
    tx4dl = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx4dl.place(x=430,y=313)
    txgkdl = Entry(cht_input, width=6 ,  font = ("Comic Sans MS",13), borderwidth=0)
    txgkdl.place(x=320,y=384)
    txckdl = Entry(cht_input, width=6 ,  font = ("Comic Sans MS",13), borderwidth=0)
    txckdl.place(x=392,y=384)         

    def diem_dia():
        save_diem(tx1dl.get(),'F2')
        save_diem(tx2dl.get(),'F3')
        save_diem(tx3dl.get(),'F4')
        save_diem(tx4dl.get(),'F5')
        save_diem(txgkdl.get(),'F6')
        save_diem(txckdl.get(),'F7')
        try:
            diem_tbdl = (float(tx1dl.get()) + float(tx2dl.get()) + float(tx3dl.get()) + float(tx4dl.get())
                        + float((txgkdl.get()))*2 + float((txckdl.get()))*3)/9
            save_diem(diem_tbdl, 'F8')
        except:
            pass

    #gdkt-pl
    tx1kt = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx1kt.place(x=549,y=313)        
    tx2kt = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx2kt.place(x=586,y=313) 
    tx3kt = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx3kt.place(x=627,y=313) 
    tx4kt = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx4kt.place(x=666,y=313)    
    txgkkt = Entry(cht_input, width=6 ,  font = ("Comic Sans MS",13), borderwidth=0)
    txgkkt.place(x=556,y=384)
    txckkt = Entry(cht_input, width=6 ,  font = ("Comic Sans MS",13), borderwidth=0)
    txckkt.place(x=627,y=384)  

    def diem_ktpl():
        save_diem(tx1kt.get(),'G2')
        save_diem(tx2kt.get(),'G3')
        save_diem(tx3kt.get(),'G4')
        save_diem(tx4kt.get(),'G5')
        save_diem(txgkkt.get(),'G6')
        save_diem(txckkt.get(),'G7')
        try:
            diem_tbkt = (float(tx1kt.get()) + float(tx2kt.get()) + float(tx3kt.get()) + float(tx4kt.get())
                        + float((txgkkt.get()))*2 + float((txckkt.get()))*3)/9
            save_diem(diem_tbkt, 'G8')
        except:
            pass

    #Tin
    tx1ti = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx1ti.place(x=775,y=313)        
    tx2ti = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx2ti.place(x=812,y=313) 
    tx3ti = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx3ti.place(x=852,y=313) 
    tx4ti = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx4ti.place(x=893,y=313)
    txgkti = Entry(cht_input, width=6 ,  font = ("Comic Sans MS",13), borderwidth=0)
    txgkti.place(x=781,y=384)
    txckti = Entry(cht_input, width=6 ,  font = ("Comic Sans MS",13), borderwidth=0)
    txckti.place(x=853,y=384) 

    def diem_tin():
        save_diem(tx1ti.get(),'H2')
        save_diem(tx2ti.get(),'H3')
        save_diem(tx3ti.get(),'H4')
        save_diem(tx4ti.get(),'H5')
        save_diem(txgkti.get(),'H6')
        save_diem(txckti.get(),'H7')
        try:
            diem_tbti = (float(tx1ti.get()) + float(tx2ti.get()) + float(tx3ti.get()) + float(tx4ti.get())
                        + float((txgkti.get()))*2 + float((txckti.get()))*3)/9
            save_diem(diem_tbti, 'H8')
        except:
            pass

    #Vật lý
    tx1vl = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx1vl.place(x=31,y=519)        
    tx2vl = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx2vl.place(x=70,y=519) 
    tx3vl = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx3vl.place(x=111,y=519) 
    tx4vl = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx4vl.place(x=150,y=519)
    txgkvl = Entry(cht_input, width=6 ,  font = ("Comic Sans MS",13), borderwidth=0)
    txgkvl.place(x=38,y=588)
    txckvl = Entry(cht_input, width=6 ,  font = ("Comic Sans MS",13), borderwidth=0)
    txckvl.place(x=109,y=588)

    def diem_ly():
        save_diem(tx1vl.get(),'I2')
        save_diem(tx2vl.get(),'I3')            
        save_diem(tx3vl.get(),'I4')
        save_diem(tx4vl.get(),'I5')
        save_diem(txgkvl.get(),'I6')
        save_diem(txckvl.get(),'I7')
        try:
            diem_tbvl = (float(tx1vl.get()) + float(tx2vl.get()) + float(tx3vl.get()) + float(tx4vl.get())
                        + float((txgkvl.get()))*2 + float((txckvl.get()))*3)/9
            save_diem(diem_tbvl, 'I8')
        except:
            pass
    #Hóa
    tx1h = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx1h.place(x=267,y=519)        
    tx2h = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx2h.place(x=304,y=519) 
    tx3h = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx3h.place(x=345,y=519) 
    tx4h = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx4h.place(x=384,y=519)
    txgkh = Entry(cht_input, width=6 ,  font = ("Comic Sans MS",13), borderwidth=0)
    txgkh.place(x=274,y=588)
    txckh = Entry(cht_input, width=6 ,  font = ("Comic Sans MS",13), borderwidth=0)
    txckh.place(x=345,y=588) 

    def diem_hoa():
        save_diem(tx1h.get(),'J2')
        save_diem(tx2h.get(),'J3')     
        save_diem(tx3h.get(),'J4')
        save_diem(tx4h.get(),'J5')
        save_diem(txgkh.get(),'J6')
        save_diem(txckh.get(),'J7')
        try:
            diem_tbh = (float(tx1h.get()) + float(tx2h.get()) + float(tx3h.get()) + float(tx4h.get())
                        + float((txgkh.get()))*2 + float((txckh.get()))*3)/9
            save_diem(diem_tbh, 'J8')
        except:
            pass

    #Sinh
    tx1s = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx1s.place(x=498,y=519)        
    tx2s = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx2s.place(x=535,y=519) 
    tx3s = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx3s.place(x=576,y=519) 
    tx4s = Entry(cht_input, width=3,  font = ("Comic Sans MS", 13), borderwidth=0)
    tx4s.place(x=615,y=519)   
    txgks = Entry(cht_input, width=6 ,  font = ("Comic Sans MS",13), borderwidth=0)
    txgks.place(x=506,y=588)
    txcks = Entry(cht_input, width=6 ,  font = ("Comic Sans MS",13), borderwidth=0)
    txcks.place(x=576,y=588)   

    def diem_sinh():
        save_diem(tx1s.get(),'K2')
        save_diem(tx2s.get(),'K3')
        save_diem(tx3s.get(),'K4')
        save_diem(tx4s.get(),'K5')
        save_diem(txgks.get(),'K6')
        save_diem(txcks.get(),'K7')
        try:
            diem_tbs = (float(tx1s.get()) + float(tx2s.get()) + float(tx3s.get()) + float(tx4s.get())
                        + float((txgks.get()))*2 + float((txcks.get()))*3)/9
            save_diem(diem_tbs, 'K8')
        except:
            pass

    def luu_diem():
        diem_toan()
        diem_van()
        diem_anh()
        diem_dia()
        diem_gdqp()
        diem_su()
        diem_ktpl()
        diem_tin()
        diem_ly()
        diem_hoa()
        diem_sinh()
        def cl_diem():
            cht_input.destroy( )
        cht_input.after(3000,cl_diem )
        vebd()
        
    save_diem_btn = Image.open(os.path.join(PATH_IMAGES, "save_diem_btn.png"))
    renderdiem = ImageTk.PhotoImage(save_diem_btn)
    imgsd = tk.Button(cht_input, image=renderdiem, borderwidth=0, command= luu_diem)
    imgsd.image = renderdiem
    imgsd.pack(side = BOTTOM, anchor= E, padx = 75, pady=75)
    
    def mo_diem():
        try:
            load_diem('A2', tx1t)
        except:
            pass
        try:
            load_diem('A3', tx2t)
        except:
            pass
        try:
            load_diem('A4', tx3t)
        except:
            pass
        try:
            load_diem('A5', tx4t)
        except:
            pass
        try:
            load_diem('A6', txgkt)
        except:
            pass
        try:
            load_diem('A7', txckt)
        except:
            pass

        
        try:
            load_diem('B2', tx1v)
        except:
            pass
        try:
            load_diem('B3', tx2v)
        except:
            pass
        try:
            load_diem('B4', tx3v)
        except:
            pass
        try:
            load_diem('B5', tx4v)
        except:
            pass
        try:
            load_diem('B6', txgkv)
        except:
            pass
        try:
            load_diem('B7', txckv)
        except:
            pass
        try:
            load_diem('C2', tx1a)
        except:
            pass
        try:
            load_diem('C3', tx2a)
        except:
            pass
        try:
            load_diem('C4', tx3a)
        except:
            pass
        try:
            load_diem('C5', tx4a)
        except:
            pass
        try:
            load_diem('C6', txgka)
        except:
            pass
        try:
            load_diem('C7', txcka)
        except:
            pass

        try:
            load_diem('D2', tx1qp)
        except:
            pass
        try:
            load_diem('D3', tx2qp)
        except:
            pass
        try:
            load_diem('D4', tx3qp)
        except:
            pass
        try:
            load_diem('D5', tx4qp)
        except:
            pass
        try:
            load_diem('D6', txgkqp)
        except:
            pass 
        try: 
            load_diem('D7', txckqp)
        except:
            pass
        try:
            load_diem('E2', tx1ls)
        except:
            pass
        try:
            load_diem('E3', tx2ls)
        except:
            pass    
        try:
            load_diem('E4', tx3ls)
        except:
            pass
        try:
            load_diem('E5', tx4ls)
        except:
            pass
        try:
            load_diem('E6', txgkls)
        except:
            pass
        try:
            load_diem('E7', txckls)
        except:
            pass

        try:
            load_diem('F2', tx1dl)
        except:
            pass
        try:
            load_diem('F3', tx2dl)
        except:
            pass    
        try:    
            load_diem('F4', tx3dl)
        except:
            pass    
        try:
            load_diem('F5', tx4dl)
        except:
            pass    
        try:
            load_diem('F6', txgkdl)
        except:
            pass    
        try:
            load_diem('F7', txckdl)
        except:
            pass

        try:
            load_diem('G2', tx1kt)
        except:
            pass
        try:
            load_diem('G3', tx2kt)
        except:
            pass
        try:
            load_diem('G4', tx3kt)
        except:
            pass
        try:
            load_diem('G5', tx4kt)
        except:
            pass
        try:
            load_diem('G6', txgkkt)
        except:
            pass
        try:
            load_diem('G7', txckkt)
        except:
            pass

        try:
            load_diem('H2', tx1ti)
        except:
            pass
        try:
            load_diem('H3', tx2ti)
        except:
            pass
        try:
            load_diem('H4', tx3ti)
        except:
            pass
        try:
            load_diem('H5', tx4ti)
        except:
            pass
        try:
            load_diem('H6', txgkti)
        except:
            pass
        try:
            load_diem('H7', txckti)
        except:
            pass

        try:
            load_diem('I2', tx1vl)
        except:
            pass
        try:
            load_diem('I3', tx2vl)
        except:
            pass
        try:
            load_diem('I4', tx3vl)
        except:
            pass
        try:
            load_diem('I5', tx4vl)
        except:
            pass
        try:
            load_diem('I6', txgkvl)
        except:
            pass
        try:
            load_diem('I7', txckvl)
        except:
            pass

        try:
            load_diem('J2', tx1h)
        except:
            pass
        try:
            load_diem('J3', tx2h)
        except:
            pass
        try:
            load_diem('J4', tx3h)
        except:
            pass
        try:
            load_diem('J5', tx4h)
        except:
            pass
        try:
            load_diem('J6', txgkh)
        except:
            pass
            
        try:
            load_diem('J7', txckh)
        except:
            pass
        try:
            load_diem('K2', tx1s)
        except:
            pass
        try:
            load_diem('K3', tx2s)
        except:
            pass
        try:
            load_diem('K4', tx3s)
        except:
            pass
        try:
            load_diem('K5', tx4s)
        except:
            pass
        try:
            load_diem('K6', txgks)
        except:
            pass
        try:
            load_diem('K7', txcks)
        except:
            pass
    mo_diem( )

    cht_input.resizable(False, False)
    cht_input.mainloop( )
vebd( )

nhap_diem_btn = Image.open(os.path.join(PATH_IMAGES, "nhapdiem.png"))
rendernd = ImageTk.PhotoImage(nhap_diem_btn)
imgnd = tk.Button(tab_cht, image=rendernd, borderwidth=0, command=cht_ip)
imgnd.image = rendernd
imgnd.pack(side = BOTTOM, anchor= E, padx = 80, pady=30)

scr.resizable(False,False )
scrn.resizable(False,False )
scrn.mainloop( )
scr.mainloop( )